#!/usr/bin/env python3
"""
POI Validation Server
Provides API endpoints for POI data validation system
Uses SQLite for data storage.
"""

from flask import Flask, jsonify, request, send_file, send_from_directory, Response
from flask_cors import CORS
import pandas as pd
import json
import os
import sqlite3
import subprocess
import ast
import requests as http_requests
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment
import trino
from trino.auth import BasicAuthentication

app = Flask(__name__)
CORS(app)

# Configuration
BASE_DIR = Path("/mnt/data/POI_Testing_Automation/version=5_0_2/test_automate_code/poi_validation_system")
DB_PATH = BASE_DIR / "poi_data.db"
SCREENSHOTS_DISPLAY = BASE_DIR / "output" / "screenshots_with_display"
SCREENSHOTS_KEPLER = BASE_DIR / "output" / "screenshots_kepler"
# Country-specific screenshot directories (India, UAE, etc.)
SCREENSHOTS_DISPLAY_IND = BASE_DIR / "output" / "IND" / "screenshots_with_display"
SCREENSHOTS_KEPLER_IND = BASE_DIR / "output" / "IND" / "screenshots_kepler"
CONFIG_FILE = BASE_DIR / "config.json"
EXCEL_OUTPUT = BASE_DIR / "output" / "poi_validation_report.xlsx"
GOOGLE_SHEETS_CSV = BASE_DIR / "output" / "poi_validation_google_sheets.csv"

# Kepler configuration
KEPLER_VITE_DIR = Path("/mnt/data/POI_Testing_Automation/version=5_0_2/get-started-vite")
KEPLER_OUTPUT_JSON = KEPLER_VITE_DIR / "src" / "output.json"
KEPLER_PORT = 8082
kepler_process = None

# Fields to return from poi_input table
INPUT_FIELDS = [
    'poi_code', 'name', 'address', 'district_code', 'country', 'latitude', 'longitude', 'gmaps_url',
    'brands', 'brand_method', 'poi_type', 'category_poi_types',
    'poi_type_cumulative', 'final_poitype_distilbert', 'final_confidence_distilbert',
    'google_category_tags', 'area_tag', 'polygon_area_sqm', 'parent_polygon_area_sqm',
    'website_domain_name'
]

# Fields to return from poi_metrics table
OUTPUT_FIELDS = [
    'name_match_pct', 'address_match_pct', 'distance_from_latlong_m',
    'location_status_match', 'ratings_diff'
]


def get_db():
    """Get a SQLite database connection"""
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    return conn


def migrate_db():
    """Add missing columns to existing database"""
    conn = get_db()
    try:
        columns = [row['name'] for row in conn.execute("PRAGMA table_info(validations)").fetchall()]
        if 'validator_name' not in columns:
            conn.execute("ALTER TABLE validations ADD COLUMN validator_name TEXT DEFAULT ''")
            conn.commit()
            print("Migration: added validator_name column to validations table")
    finally:
        conn.close()


def safe_parse_list(value):
    """Safely parse list strings from DB"""
    if value is None or value == '' or value == 'nan':
        return []
    if isinstance(value, str):
        try:
            if value.startswith('[') and value.endswith(']'):
                return ast.literal_eval(value)
            return [value]
        except:
            return [value] if value else []
    return value if isinstance(value, list) else [value]


def find_screenshot(poi_code, screenshot_dir):
    """Find screenshot file for given POI code, checking country-specific dirs too"""
    pattern = f"{poi_code}*.png"
    files = list(screenshot_dir.glob(pattern))
    if files:
        return files[0].name
    # Check country-specific directories (IND, etc.)
    country_dir = screenshot_dir.parent / "IND" / screenshot_dir.name
    if country_dir.exists():
        files = list(country_dir.glob(pattern))
        if files:
            return files[0].name
    return None


def load_config():
    """Load config from JSON file"""
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, 'r') as f:
                return json.load(f)
        except:
            return {"assignees": []}
    return {"assignees": []}


@app.route('/')
def index():
    """Serve the main HTML page"""
    return send_file(BASE_DIR / 'poi_validation_app.html')


@app.route('/logo.svg')
def serve_logo():
    """Serve the Sherlock logo SVG"""
    return send_file(BASE_DIR / 'sherlockLogo.svg', mimetype='image/svg+xml')


@app.route('/mycroft_logo.png')
def serve_mycroft_logo():
    """Serve the Mycroft logo PNG"""
    return send_file(BASE_DIR / 'MyCroft_logo.png', mimetype='image/png')

@app.route('/favicon.png')
def serve_favicon():
    """Serve the favicon"""
    return send_file(BASE_DIR / 'MyCroft_logo.png', mimetype='image/png')


@app.route('/api/config')
def get_config():
    """Get configuration including assignee list"""
    return jsonify(load_config())


@app.route('/api/poi_type_descriptions')
def get_poi_type_descriptions():
    """Get all POI type descriptions from CSV"""
    csv_path = BASE_DIR / "poi_type_dexcription.csv"
    if not csv_path.exists():
        return jsonify({'descriptions': {}})
    try:
        df = pd.read_csv(csv_path)
        descriptions = dict(zip(df['poi_type'].astype(str), df['description'].astype(str)))
        return jsonify({'descriptions': descriptions})
    except Exception as e:
        print(f"Error reading poi_type descriptions: {e}")
        return jsonify({'descriptions': {}})


@app.route('/api/filters')
def get_filters():
    """Get unique values for filter dropdowns"""
    conn = get_db()
    try:
        countries = [r[0] for r in conn.execute(
            "SELECT DISTINCT country FROM poi_input WHERE country IS NOT NULL ORDER BY country"
        ).fetchall()]

        poi_types = [r[0] for r in conn.execute(
            "SELECT DISTINCT poi_type FROM poi_input WHERE poi_type IS NOT NULL ORDER BY poi_type"
        ).fetchall()]

        # Extract level1 from poi_type
        level1_raw = conn.execute(
            "SELECT DISTINCT poi_type FROM poi_input WHERE poi_type IS NOT NULL"
        ).fetchall()
        level1_values = sorted(set(r[0].split('.')[0] for r in level1_raw if r[0]))

        return jsonify({
            'countries': countries,
            'poi_types': poi_types,
            'level1_values': level1_values
        })
    finally:
        conn.close()


@app.route('/api/poi_list')
def get_poi_list():
    """Get list of all POIs with optional filtering"""
    country = request.args.get('country', '')
    poi_type = request.args.get('poi_type', '')
    level1 = request.args.get('level1', '')
    assignee = request.args.get('assignee', '')

    conn = get_db()
    try:
        # First get ALL POIs sorted by poi_code (for consistent round-robin)
        all_rows = conn.execute("SELECT poi_code, name, poi_type, country FROM poi_input ORDER BY poi_code").fetchall()
        all_pois = [{'poi_code': r['poi_code'], 'name': r['name'] or 'Unknown',
                     'poi_type': r['poi_type'] or '', 'country': r['country'] or ''} for r in all_rows]

        # Apply assignee round-robin FIRST (on the full list)
        if assignee:
            config = load_config()
            assignees = config.get('assignees', [])
            if assignees and assignee in assignees:
                assignee_index = assignees.index(assignee)
                num_assignees = len(assignees)
                all_pois = [p for i, p in enumerate(all_pois) if i % num_assignees == assignee_index]

        # Then apply filters on the assignee's fixed set
        poi_list = all_pois
        if country:
            poi_list = [p for p in poi_list if p['country'] == country]
        if level1:
            poi_list = [p for p in poi_list if p['poi_type'].startswith(f"{level1}.")]
        if poi_type:
            poi_list = [p for p in poi_list if p['poi_type'] == poi_type]

        # Strip extra fields before returning
        poi_list = [{'poi_code': p['poi_code'], 'name': p['name']} for p in poi_list]

        return jsonify({'poi_list': poi_list})
    finally:
        conn.close()


@app.route('/api/poi_data/<poi_code>')
def get_poi_data(poi_code):
    """Get detailed data for a specific POI"""
    conn = get_db()
    try:
        # Get input data
        row = conn.execute("SELECT * FROM poi_input WHERE poi_code = ?", (poi_code,)).fetchone()
        if not row:
            return jsonify({'error': f'POI {poi_code} not found'}), 404

        data = {'poi_code': poi_code}

        # Add input fields
        for field in INPUT_FIELDS:
            if field in row.keys():
                value = row[field]
                if field in ['category_poi_types', 'google_category_tags']:
                    data[field] = safe_parse_list(value)
                else:
                    data[field] = value

        # Get metrics data
        metrics = conn.execute("SELECT * FROM poi_metrics WHERE poi_code = ?", (poi_code,)).fetchone()
        if metrics:
            for field in OUTPUT_FIELDS:
                if field in metrics.keys():
                    data[field] = metrics[field]

        # Find screenshots
        screenshot_display = find_screenshot(poi_code, SCREENSHOTS_DISPLAY)
        screenshot_kepler = find_screenshot(poi_code, SCREENSHOTS_KEPLER)

        if screenshot_display:
            data['screenshot_display'] = f"display/{screenshot_display}"
        if screenshot_kepler:
            data['screenshot_kepler'] = f"kepler/{screenshot_kepler}"

        return jsonify(data)
    finally:
        conn.close()


@app.route('/api/screenshot/display/<path:filename>')
def serve_screenshot_display(filename):
    """Serve screenshot files from display directory (checks USA then IND)"""
    if (SCREENSHOTS_DISPLAY / filename).exists():
        return send_from_directory(SCREENSHOTS_DISPLAY, filename)
    elif (SCREENSHOTS_DISPLAY_IND / filename).exists():
        return send_from_directory(SCREENSHOTS_DISPLAY_IND, filename)
    else:
        return jsonify({'error': 'Display screenshot not found'}), 404


@app.route('/api/screenshot/kepler/<path:filename>')
def serve_screenshot_kepler(filename):
    """Serve screenshot files from kepler directory (checks USA then IND)"""
    if (SCREENSHOTS_KEPLER / filename).exists():
        return send_from_directory(SCREENSHOTS_KEPLER, filename)
    elif (SCREENSHOTS_KEPLER_IND / filename).exists():
        return send_from_directory(SCREENSHOTS_KEPLER_IND, filename)
    else:
        return jsonify({'error': 'Kepler screenshot not found'}), 404


@app.route('/api/get_validations')
def get_validations():
    """Get all existing validations"""
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM validations").fetchall()
        validations = {}
        for r in rows:
            validations[r['poi_code']] = {
                'poi_code': r['poi_code'],
                'poi_type_validation': r['poi_type_validation'] or '',
                'correct_poi_type': r['correct_poi_type'] or '',
                'brand_validation': r['brand_validation'] or '',
                'polygon_area_validation': r['polygon_area_validation'] or '',
                'polygon_validation': r['polygon_validation'] or '',
                'comments': r['comments'] or '',
                'timestamp': r['timestamp'] or '',
                'validator_name': r['validator_name'] or ''
            }
        return jsonify(validations)
    finally:
        conn.close()


@app.route('/api/save_validation', methods=['POST'])
def save_validation():
    """Save validation for a POI"""
    try:
        validation_data = request.json
        poi_code = validation_data.get('poi_code')

        if not poi_code:
            return jsonify({'error': 'POI code is required'}), 400

        conn = get_db()
        try:
            conn.execute("""
                INSERT OR REPLACE INTO validations
                (poi_code, poi_type_validation, correct_poi_type, brand_validation,
                 polygon_area_validation, polygon_validation, comments, timestamp, validator_name)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                poi_code,
                validation_data.get('poi_type_validation', ''),
                validation_data.get('correct_poi_type', ''),
                validation_data.get('brand_validation', ''),
                validation_data.get('polygon_area_validation', ''),
                validation_data.get('polygon_validation', ''),
                validation_data.get('comments', ''),
                validation_data.get('timestamp', ''),
                validation_data.get('validator_name', '')
            ))
            conn.commit()
        finally:
            conn.close()

        # Update exports
        print(f"Auto-saving validation for {poi_code} to Excel...")
        update_excel_report()
        update_google_sheets_csv()

        return jsonify({
            'success': True,
            'message': f'Validation saved and Excel updated for {poi_code}'
        })
    except Exception as e:
        print(f"Error saving validation: {e}")
        return jsonify({'error': str(e)}), 500


def get_merged_dataframe():
    """Get a merged DataFrame of poi_input + poi_metrics for export purposes"""
    conn = get_db()
    try:
        df = pd.read_sql_query("""
            SELECT i.*, m.name_match_pct, m.address_match_pct,
                   m.distance_from_latlong_m, m.location_status_match, m.ratings_diff
            FROM poi_input i
            LEFT JOIN poi_metrics m ON i.poi_code = m.poi_code
        """, conn)
        return df
    finally:
        conn.close()


def get_all_validations_dict():
    """Get all validations as a dictionary"""
    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM validations").fetchall()
        return {r['poi_code']: dict(r) for r in rows}
    finally:
        conn.close()


def update_excel_report():
    """Update Excel report with all POI data and validations"""
    try:
        df = get_merged_dataframe()
        if df.empty:
            return

        validations = get_all_validations_dict()

        # Add validation columns
        df['poi_type_validation'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('poi_type_validation', ''))
        df['correct_poi_type'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('correct_poi_type', ''))
        df['brand_validation'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('brand_validation', ''))
        df['polygon_area_validation'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('polygon_area_validation', ''))
        df['polygon_validation'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('polygon_validation', ''))
        df['validation_comments'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('comments', ''))
        df['validation_timestamp'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('timestamp', ''))
        df['validator_name'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('validator_name', ''))

        with pd.ExcelWriter(EXCEL_OUTPUT, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='POI Validation Data', index=False)

            workbook = writer.book
            worksheet = workbook['POI Validation Data']

            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)

            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            validation_fill_correct = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            validation_fill_incorrect = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

            validation_cols = ['poi_type_validation', 'brand_validation', 'polygon_area_validation', 'polygon_validation']
            for col_name in validation_cols:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col_idx)
                        if cell.value == 'correct':
                            cell.fill = validation_fill_correct
                        elif cell.value == 'incorrect':
                            cell.fill = validation_fill_incorrect

            summary_data = {
                'Metric': [
                    'Total POIs', 'POIs Validated',
                    'POI Type - Correct', 'POI Type - Incorrect',
                    'Brand - Correct', 'Brand - Incorrect',
                    'Polygon Area - Correct', 'Polygon Area - Incorrect',
                    'Polygon - Correct', 'Polygon - Incorrect'
                ],
                'Count': [
                    len(df),
                    len([v for v in validations.values() if any([v.get('poi_type_validation'), v.get('brand_validation'), v.get('polygon_area_validation'), v.get('polygon_validation')])]),
                    len([v for v in validations.values() if v.get('poi_type_validation') == 'correct']),
                    len([v for v in validations.values() if v.get('poi_type_validation') == 'incorrect']),
                    len([v for v in validations.values() if v.get('brand_validation') == 'correct']),
                    len([v for v in validations.values() if v.get('brand_validation') == 'incorrect']),
                    len([v for v in validations.values() if v.get('polygon_area_validation') == 'correct']),
                    len([v for v in validations.values() if v.get('polygon_area_validation') == 'incorrect']),
                    len([v for v in validations.values() if v.get('polygon_validation') == 'correct']),
                    len([v for v in validations.values() if v.get('polygon_validation') == 'incorrect'])
                ]
            }

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Validation Summary', index=False)

            summary_sheet = workbook['Validation Summary']
            for col in range(1, 3):
                cell = summary_sheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            summary_sheet.column_dimensions['A'].width = 25
            summary_sheet.column_dimensions['B'].width = 15

    except Exception as e:
        print(f"Error updating Excel report: {e}")


def update_google_sheets_csv():
    """Create/update Google Sheets compatible CSV with validation data"""
    try:
        df = get_merged_dataframe()
        if df.empty:
            return

        validations = get_all_validations_dict()

        df['poi_type_validation'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('poi_type_validation', ''))
        df['correct_poi_type'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('correct_poi_type', ''))
        df['brand_validation'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('brand_validation', ''))
        df['polygon_area_validation'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('polygon_area_validation', ''))
        df['polygon_validation'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('polygon_validation', ''))
        df['validation_comments'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('comments', ''))
        df['validation_timestamp'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('timestamp', ''))
        df['validator_name'] = df['poi_code'].map(lambda x: validations.get(x, {}).get('validator_name', ''))

        google_sheets_columns = [
            'poi_code', 'name', 'address', 'district_code', 'country',
            'poi_type', 'final_poitype_distilbert', 'final_confidence_distilbert',
            'polygon_area_sqm', 'area_tag',
            'name_match_pct', 'address_match_pct', 'distance_from_latlong_m', 'location_status_match',
            'poi_type_validation', 'correct_poi_type', 'brand_validation',
            'polygon_area_validation', 'polygon_validation',
            'validation_comments', 'validation_timestamp', 'validator_name'
        ]

        available_columns = [col for col in google_sheets_columns if col in df.columns]
        google_sheets_df = df[available_columns].copy()

        for col in google_sheets_df.columns:
            if google_sheets_df[col].dtype == 'object':
                google_sheets_df[col] = google_sheets_df[col].astype(str).replace(['nan', 'None'], '')

        google_sheets_df.to_csv(GOOGLE_SHEETS_CSV, index=False)
        print(f"Google Sheets CSV updated: {GOOGLE_SHEETS_CSV}")

    except Exception as e:
        print(f"Error updating Google Sheets CSV: {e}")


@app.route('/api/analytics')
def get_analytics():
    """Get analytics data for the dashboard"""
    conn = get_db()
    try:
        # Get all POI codes and names
        poi_rows = conn.execute("SELECT poi_code, name FROM poi_input ORDER BY poi_code").fetchall()
        if not poi_rows:
            return jsonify({'error': 'No data available'}), 404

        all_poi_codes = [r['poi_code'] for r in poi_rows]
        all_poi_names = {r['poi_code']: r['name'] or 'Unknown' for r in poi_rows}

        # Get all validations
        val_rows = conn.execute("SELECT * FROM validations").fetchall()
        validations = {r['poi_code']: dict(r) for r in val_rows}

        config = load_config()
        assignees = config.get('assignees', [])

        def compute_stats(poi_codes):
            total = len(poi_codes)
            tested = 0
            pois_with_incorrect = 0
            stats = {
                'total': total,
                'poi_type': {'correct': 0, 'incorrect': 0, 'untested': 0},
                'brand': {'correct': 0, 'incorrect': 0, 'untested': 0},
                'polygon_area': {'correct': 0, 'incorrect': 0, 'untested': 0},
                'polygon': {'correct': 0, 'incorrect': 0, 'untested': 0},
                'comments': []
            }
            for pc in poi_codes:
                v = validations.get(pc, {})
                has_any = bool(v.get('poi_type_validation') or v.get('brand_validation')
                              or v.get('polygon_area_validation') or v.get('polygon_validation'))
                if has_any:
                    tested += 1

                has_incorrect = False
                for key, field in [('poi_type', 'poi_type_validation'), ('brand', 'brand_validation'),
                                   ('polygon_area', 'polygon_area_validation'), ('polygon', 'polygon_validation')]:
                    val = v.get(field, '')
                    if val == 'correct':
                        stats[key]['correct'] += 1
                    elif val == 'incorrect':
                        stats[key]['incorrect'] += 1
                        has_incorrect = True
                    else:
                        stats[key]['untested'] += 1

                if has_incorrect:
                    pois_with_incorrect += 1

                comment = (v.get('comments') or '').strip()
                if comment:
                    stats['comments'].append({
                        'poi_code': pc,
                        'name': all_poi_names.get(pc, 'Unknown'),
                        'comment': comment,
                        'timestamp': v.get('timestamp', ''),
                        'correct_poi_type': v.get('correct_poi_type', '')
                    })

            stats['tested'] = tested
            stats['untested_count'] = total - tested
            stats['total_incorrect_pois'] = pois_with_incorrect
            return stats

        overall = compute_stats(all_poi_codes)

        per_assignee = {}
        if assignees:
            for assignee in assignees:
                assignee_codes = [pc for pc in all_poi_codes
                                  if validations.get(pc, {}).get('validator_name', '') == assignee]
                per_assignee[assignee] = compute_stats(assignee_codes)

        # Tester analytics: per-validator stats based on actual validator_name
        from datetime import datetime
        tester_stats = []
        # Group validations by validator_name
        validators_map = {}
        for pc, v in validations.items():
            vname = v.get('validator_name', '')
            if not vname:
                continue
            if vname not in validators_map:
                validators_map[vname] = []
            validators_map[vname].append(v)

        for vname in sorted(validators_map.keys()):
            v_list = validators_map[vname]
            pois_tested = len(v_list)

            poi_type_correct = sum(1 for v in v_list if v.get('poi_type_validation') == 'correct')
            poi_type_incorrect = sum(1 for v in v_list if v.get('poi_type_validation') == 'incorrect')
            brand_correct = sum(1 for v in v_list if v.get('brand_validation') == 'correct')
            brand_incorrect = sum(1 for v in v_list if v.get('brand_validation') == 'incorrect')
            polygon_area_correct = sum(1 for v in v_list if v.get('polygon_area_validation') == 'correct')
            polygon_area_incorrect = sum(1 for v in v_list if v.get('polygon_area_validation') == 'incorrect')
            polygon_correct = sum(1 for v in v_list if v.get('polygon_validation') == 'correct')
            polygon_incorrect = sum(1 for v in v_list if v.get('polygon_validation') == 'incorrect')

            # POI-level correct/incorrect
            all_correct = 0
            any_incorrect = 0
            for v in v_list:
                fields = [v.get('poi_type_validation', ''), v.get('brand_validation', ''),
                          v.get('polygon_area_validation', ''), v.get('polygon_validation', '')]
                filled = [f for f in fields if f]
                if filled and all(f == 'correct' for f in filled):
                    all_correct += 1
                if any(f == 'incorrect' for f in fields):
                    any_incorrect += 1

            # Timestamps
            timestamps = []
            for v in v_list:
                ts = v.get('timestamp', '')
                if ts:
                    try:
                        timestamps.append(datetime.fromisoformat(ts.replace('Z', '+00:00')))
                    except:
                        pass

            start_time = ''
            end_time = ''
            avg_seconds = 0
            if timestamps:
                timestamps.sort()
                start_time = timestamps[0].strftime('%Y-%m-%d %H:%M:%S UTC')
                end_time = timestamps[-1].strftime('%Y-%m-%d %H:%M:%S UTC')
                if pois_tested > 1:
                    total_seconds = (timestamps[-1] - timestamps[0]).total_seconds()
                    avg_seconds = round(total_seconds / (pois_tested - 1))

            tester_stats.append({
                'name': vname,
                'pois_tested': pois_tested,
                'all_correct': all_correct,
                'any_incorrect': any_incorrect,
                'poi_type': {'correct': poi_type_correct, 'incorrect': poi_type_incorrect},
                'brand': {'correct': brand_correct, 'incorrect': brand_incorrect},
                'polygon_area': {'correct': polygon_area_correct, 'incorrect': polygon_area_incorrect},
                'polygon': {'correct': polygon_correct, 'incorrect': polygon_incorrect},
                'start_time': start_time,
                'end_time': end_time,
                'avg_seconds_per_poi': avg_seconds
            })

        return jsonify({
            'overall': overall,
            'per_assignee': per_assignee,
            'assignees': assignees,
            'tester_stats': sorted(tester_stats, key=lambda t: t['pois_tested'], reverse=True)
        })
    finally:
        conn.close()


@app.route('/api/summary')
def get_summary():
    """Get summary data: which poi_types have the most incorrect validations"""
    conn = get_db()
    try:
        # Join poi_input with validations to get poi_type per validation
        rows = conn.execute("""
            SELECT i.poi_type, v.poi_type_validation, v.brand_validation,
                   v.polygon_area_validation, v.polygon_validation
            FROM poi_input i
            INNER JOIN validations v ON i.poi_code = v.poi_code
            WHERE i.poi_type IS NOT NULL
        """).fetchall()

        # Count incorrects per poi_type for each validation field
        poi_type_incorrect = {}
        brand_incorrect = {}
        polygon_area_incorrect = {}
        polygon_incorrect = {}

        # Also count totals per poi_type for context
        poi_type_totals = {}

        for r in rows:
            pt = r['poi_type']
            poi_type_totals[pt] = poi_type_totals.get(pt, 0) + 1

            if r['poi_type_validation'] == 'incorrect':
                poi_type_incorrect[pt] = poi_type_incorrect.get(pt, 0) + 1
            if r['brand_validation'] == 'incorrect':
                brand_incorrect[pt] = brand_incorrect.get(pt, 0) + 1
            if r['polygon_area_validation'] == 'incorrect':
                polygon_area_incorrect[pt] = polygon_area_incorrect.get(pt, 0) + 1
            if r['polygon_validation'] == 'incorrect':
                polygon_incorrect[pt] = polygon_incorrect.get(pt, 0) + 1

        def to_ranked_list(counts_dict):
            return sorted(
                [{'poi_type': k, 'incorrect': v, 'total_validated': poi_type_totals.get(k, 0)}
                 for k, v in counts_dict.items()],
                key=lambda x: x['incorrect'], reverse=True
            )

        return jsonify({
            'poi_type_validation': to_ranked_list(poi_type_incorrect),
            'brand_validation': to_ranked_list(brand_incorrect),
            'polygon_area_validation': to_ranked_list(polygon_area_incorrect),
            'polygon_validation': to_ranked_list(polygon_incorrect)
        })
    finally:
        conn.close()


@app.route('/api/filtered_pois')
def get_filtered_pois():
    """Get POIs filtered by a specific validation field and value"""
    field = request.args.get('field', '')
    value = request.args.get('value', '')
    assignee = request.args.get('assignee', '')

    valid_fields = {
        'poi_type_validation': 'poi_type_validation',
        'brand_validation': 'brand_validation',
        'polygon_area_validation': 'polygon_area_validation',
        'polygon_validation': 'polygon_validation'
    }

    if field not in valid_fields:
        return jsonify({'error': f'Invalid field: {field}'}), 400

    conn = get_db()
    try:
        # Get all POIs with names
        all_pois = conn.execute("SELECT poi_code, name FROM poi_input ORDER BY poi_code").fetchall()
        poi_list = [{'poi_code': r['poi_code'], 'name': r['name'] or 'Unknown'} for r in all_pois]

        # Get validations
        val_rows = conn.execute("SELECT * FROM validations").fetchall()
        validations = {r['poi_code']: dict(r) for r in val_rows}

        # Filter by validation field + value, and optionally by assignee
        filtered = []
        for p in poi_list:
            v = validations.get(p['poi_code'], {})

            # Filter by assignee (actual validator)
            if assignee and v.get('validator_name', '') != assignee:
                continue

            val = v.get(field, '')
            if value == 'untested':
                if not val:
                    filtered.append(p)
            elif val == value:
                filtered.append(p)

        return jsonify({'poi_list': filtered})
    finally:
        conn.close()


def fetch_poi_polygon_from_trino(poi_code):
    """Query Trino for a single POI's polygon data"""
    conn = trino.dbapi.connect(
        host='prestoazure.infiniteanalytics.com',
        port=443,
        user='application',
        catalog='delta',
        schema='default',
        auth=BasicAuthentication('application', 'knob#or53RainEin5teinTom168'),
        http_scheme='https',
        client_tags=['test']
    )
    cursor = conn.cursor()
    query = f"""
    SELECT poi_code, polygon, latitude, longitude, name, address
    FROM poi_data_5_0_10
    WHERE poi_code = '{poi_code}'
    """
    cursor.execute(query)
    results = cursor.fetchall()
    column_names = [desc[0] for desc in cursor.description]
    cursor.close()
    conn.close()

    if not results:
        return None

    df = pd.DataFrame(results, columns=column_names)
    return df.to_dict(orient="records")


def is_kepler_running():
    """Check if the Kepler Vite dev server is running on the expected port"""
    import socket
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        return s.connect_ex(('localhost', KEPLER_PORT)) == 0


def start_kepler_server():
    """Start the Kepler Vite dev server if not already running"""
    global kepler_process
    if is_kepler_running():
        print(f"Kepler dev server already running on port {KEPLER_PORT}")
        return True

    try:
        print(f"Starting Kepler dev server in {KEPLER_VITE_DIR}...")
        kepler_process = subprocess.Popen(
            ['pnpm', 'dev', '--host'],
            cwd=str(KEPLER_VITE_DIR),
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            preexec_fn=os.setsid
        )
        import time
        for _ in range(30):
            time.sleep(1)
            if is_kepler_running():
                print(f"Kepler dev server started on port {KEPLER_PORT}")
                return True
        print("Kepler dev server failed to start within 30 seconds")
        return False
    except Exception as e:
        print(f"Error starting Kepler dev server: {e}")
        return False


@app.route('/api/open_kepler/<poi_code>', methods=['POST'])
def open_kepler(poi_code):
    """Fetch polygon data for a POI from Trino, write to output.json, and start Kepler"""
    try:
        print(f"Fetching polygon data for {poi_code} from Trino...")
        poi_data = fetch_poi_polygon_from_trino(poi_code)

        if not poi_data:
            return jsonify({'error': f'No polygon data found for {poi_code} in Trino'}), 404

        with open(KEPLER_OUTPUT_JSON, 'w') as f:
            json.dump(poi_data, f, indent=2)
        print(f"Wrote {len(poi_data)} records to {KEPLER_OUTPUT_JSON}")

        kepler_started = start_kepler_server()

        if not kepler_started:
            return jsonify({
                'error': 'Could not start Kepler dev server. Please start it manually: cd get-started-vite && pnpm dev'
            }), 500

        return jsonify({
            'success': True,
            'message': f'Kepler ready for {poi_code}',
            'kepler_url': '/kepler/'
        })

    except Exception as e:
        print(f"Error opening Kepler for {poi_code}: {e}")
        return jsonify({'error': str(e)}), 500


def _proxy_to_kepler(path=''):
    """Proxy a request to the Kepler Vite dev server"""
    target_url = f"http://localhost:{KEPLER_PORT}/{path}"
    if request.query_string:
        target_url += f"?{request.query_string.decode()}"

    try:
        resp = http_requests.request(
            method=request.method,
            url=target_url,
            headers={k: v for k, v in request.headers if k.lower() != 'host'},
            data=request.get_data(),
            allow_redirects=False,
            timeout=30
        )

        content = resp.content
        content_type = resp.headers.get('Content-Type', '')

        excluded_headers = ['content-encoding', 'content-length', 'transfer-encoding', 'connection']
        headers = {k: v for k, v in resp.headers.items() if k.lower() not in excluded_headers}

        return Response(content, status=resp.status_code, headers=headers, content_type=content_type)

    except http_requests.ConnectionError:
        return jsonify({'error': 'Kepler dev server is not running. Click "View Live Kepler" on a POI first.'}), 502
    except Exception as e:
        return jsonify({'error': f'Proxy error: {str(e)}'}), 500


# Kepler proxy — main entry point
@app.route('/kepler/')
def proxy_kepler_root():
    return _proxy_to_kepler('')

# Proxy all Vite internal paths that Kepler needs
@app.route('/@vite/<path:path>')
def proxy_vite_client(path):
    return _proxy_to_kepler(f'@vite/{path}')

@app.route('/@react-refresh')
def proxy_react_refresh():
    return _proxy_to_kepler('@react-refresh')

@app.route('/src/<path:path>')
def proxy_src(path):
    return _proxy_to_kepler(f'src/{path}')

@app.route('/node_modules/<path:path>')
def proxy_node_modules(path):
    return _proxy_to_kepler(f'node_modules/{path}')


if __name__ == '__main__':
    (BASE_DIR / "output").mkdir(exist_ok=True)
    migrate_db()

    print(f"Starting POI Validation Server...")
    print(f"Base directory: {BASE_DIR}")
    print(f"Database: {DB_PATH}")
    print(f"Screenshots Display: {SCREENSHOTS_DISPLAY}")
    print(f"Screenshots Kepler: {SCREENSHOTS_KEPLER}")
    print(f"\nServer running at http://localhost:5002")
    print("Open your browser and navigate to http://localhost:5002 to use the application")

    app.run(debug=True, host='0.0.0.0', port=5002)
